from django.contrib import messages
from django.forms import inlineformset_factory
from django.http import JsonResponse, HttpResponse
from django.template.loader import get_template
from django.views.generic.edit import FormView
from django.core.paginator import Paginator

from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse

from openpyxl.utils import get_column_letter
from openpyxl.workbook import Workbook
from openpyxl.worksheet.dimensions import ColumnDimension
from openpyxl.styles import Font
from xhtml2pdf import pisa

from .forms import CadastroClienteForm, VeiculoForm, CriarOrcamentoForm
from .models import Cliente, Veiculo, OrcamentoNaoAprovado, OrdemServicoAberta, OrdemServicoConcluida
from django.db.models import F, Count


def clientes(request):
    form = CadastroClienteForm()
    clientes = Cliente.objects.all()  # Obtenha todos os clientes

    clientes_com_veiculo = []
    for cliente in clientes:
        veiculos = Veiculo.objects.filter(cliente=cliente)
        if veiculos.exists():
            clientes_com_veiculo.append({
                'cliente': cliente,
                'veiculos': veiculos
            })

    paginator = Paginator(clientes_com_veiculo, 3)  # Dividir em páginas de 10 itens cada
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    qtd_clientes = len(clientes_com_veiculo)

    return render(request, 'pages/clientes.html', {'form': form,
                                                   'page_obj': page_obj,  # Adicionar a página atual ao contexto
                                                   'qtd_clientes': qtd_clientes})


def clientes_lista(request):
    form = CadastroClienteForm()
    queryset = Cliente.objects.select_related('veiculo').values('id',
                                                                'nome',
                                                                'cpf',
                                                                'endereco',
                                                                'telefone',
                                                                'email',
                                                                'veiculo__marca',
                                                                'veiculo__modelo',
                                                                'veiculo__placa',
                                                                'veiculo__cor',  # Inclua o campo 'cor' do veículo aqui
                                                                'veiculo__km'    # Inclua o campo 'km' do veículo aqui
                                                                )
    result_list = list(queryset)

    clientes = []

    for item in result_list:
        if item['veiculo__placa']:
            clientes.append(item)

    return JsonResponse({
        'clientes': clientes,
    })


def cadastro_clientes(request):
    if request.method == "GET":
        form = CadastroClienteForm()
        form_veiculo_factory = inlineformset_factory(Cliente, Veiculo,
                                                     form=VeiculoForm, extra=1)
        form_veiculo = form_veiculo_factory()
        context = {
            'form': form,
            'form_veiculo': form_veiculo,
        }
        return render(request, "pages/cadastro_clientes.html", context)
    elif request.method == "POST":
        form = CadastroClienteForm(request.POST)
        form_veiculo_factory = inlineformset_factory(Cliente, Veiculo,
                                                     form=VeiculoForm)
        form_veiculo = form_veiculo_factory(request.POST)
        if form.is_valid() and form_veiculo.is_valid():
            cliente = form.save()
            form_veiculo.instance = cliente
            form_veiculo.save()
            return redirect(reverse('clientes'))
        else:
            context = {
                'form': form,
                'form_veiculo': form_veiculo,
            }
            return render(request, "pages/cadastro_clientes.html", context)


def informacoes_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, id=cliente_id)
    veiculos = Veiculo.objects.filter(cliente=cliente)

    cliente_detalhes = {
        'nome': cliente.nome,
        'cpf': cliente.cpf,
        'telefone': cliente.telefone,
        'email': cliente.email,
        'veiculos': [{'modelo': veiculo.modelo,
                      'placa': veiculo.placa,
                      'ano': veiculo.ano,
                      'km': veiculo.km,
                      'cor': veiculo.cor} for veiculo in veiculos]
    }

    return JsonResponse(cliente_detalhes)

def editar(request, cliente_id):
    if request.method == "GET":
        objeto = Cliente.objects.filter(id=cliente_id).first()
        if objeto is None:
            return redirect(reverse('clientes'))
        form = CadastroClienteForm(instance=objeto)
        form_veiculo_factory = inlineformset_factory(Cliente, Veiculo, form=VeiculoForm, extra=0)
        form_veiculo = form_veiculo_factory(instance=objeto)
        context = {
            'form': form,
            'form_veiculo': form_veiculo,
        }
        return render(request, "pages/cadastro_clientes.html", context)
    elif request.method == "POST":
        objeto = Cliente.objects.filter(id=cliente_id).first()
        if objeto is None:
            return redirect(reverse('clientes'))
        form = CadastroClienteForm(request.POST, instance=objeto)
        form_veiculo_factory = inlineformset_factory(Cliente, Veiculo, form=VeiculoForm)
        form_veiculo = form_veiculo_factory(request.POST, instance=objeto)
        if form.is_valid() and form_veiculo.is_valid():
            cliente = form.save()
            form_veiculo.instance = cliente
            form_veiculo.save()
            return redirect(reverse('clientes'))
        else:
            context = {
                'form': form,
                'form_veiculo': form_veiculo,
            }
            return render(request, "pages/cadastro_clientes.html", context)


def excluir_cliente(request, cliente_id):
    cliente = get_object_or_404(Cliente, pk=cliente_id)
    cliente.delete()

    id_cliente = cliente_id
    return redirect(reverse('clientes') + f'?aba=att_cliente&id_cliente={id_cliente}')


def exportar_clientes(request):
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="clientes_com_carro.xlsx"'

    wb = Workbook()
    ws = wb.active

    row_num = 1  # Comece da primeira linha, pois a primeira linha contém os títulos das colunas

    bold_font = Font(bold=True)
    bold_style = "bold"  # Não é necessário criar um estilo para negrito

    columns = ['Nome', 'CPF', 'Telefone', 'Email', 'Marca do Carro', 'Modelo do Carro', 'Placa do Carro']

    # Crie uma lista para rastrear a largura máxima de cada coluna
    column_widths = [len(column_title) for column_title in columns]

    for col_num, column_title in enumerate(columns, 1):
        cell = ws.cell(row=row_num, column=col_num, value=column_title)
        cell.font = bold_font
        # Atualize a largura máxima da coluna com base no título
        column_widths[col_num - 1] = max(column_widths[col_num - 1], len(column_title))

    clientes_com_carro = Cliente.objects.filter(veiculo__isnull=False).distinct()

    for cliente in clientes_com_carro:
        veiculos = Veiculo.objects.filter(cliente=cliente)
        for veiculo in veiculos:
            row_num += 1
            row = [
                cliente.nome,
                cliente.cpf,
                cliente.telefone,
                cliente.email,
                veiculo.marca,
                veiculo.modelo,
                veiculo.placa
            ]
            for col_num, cell_value in enumerate(row, 1):
                cell = ws.cell(row=row_num, column=col_num, value=cell_value)
                # Atualize a largura máxima da coluna com base no conteúdo da célula
                column_widths[col_num - 1] = max(column_widths[col_num - 1], len(str(cell_value)))

    # Ajuste a largura de cada coluna de acordo com a largura máxima
    for col_num, width in enumerate(column_widths, 1):
        column_letter = get_column_letter(col_num)
        ws.column_dimensions[column_letter].width = (width + 2)  # Adicione um espaço extra

    wb.save(response)
    return response


def novo_orcamento(request):
    queryset = Cliente.objects.select_related('veiculo').values(
        'id',
        'nome',
        'cpf',
        'endereco',
        'telefone',
        'email',
        'veiculo__marca',
        'veiculo__modelo',
        'veiculo__placa',
        'veiculo__cor',
        'veiculo__km'
    )

    clientes = [item for item in queryset if item['veiculo__placa']]

    paginator = Paginator(clientes, 10)  # Dividir em páginas de 10 itens cada
    page_number = request.GET.get('page')  # Obter o número da página da query string
    page_obj = paginator.get_page(page_number)

    return render(request, 'pages/novo_orcamento.html', {'page_obj': page_obj} )


def contar_orcamentos_nao_aprovados(request):
    clientes_com_orcamentos = Cliente.objects.filter(
        orcamentonaoaprovado__isnull=False
    ).annotate(
        num_orcamentos_nao_aprovados=Count('orcamentonaoaprovado')
    ).values(
        'id',
        'nome',
        'veiculo__marca',
        'veiculo__modelo',
        'veiculo__placa',
        'num_orcamentos_nao_aprovados'
    )
    return JsonResponse({'clientes': clientes})

def criar_orcamento(request, cliente_id):
    cliente = get_object_or_404(Cliente, pk=cliente_id)
    veiculo = Veiculo.objects.filter(cliente=cliente).first()  # Obtém o primeiro veículo associado ao cliente

    if request.method == 'POST':
        form = CriarOrcamentoForm(request.POST)
        if form.is_valid():
            orcamento = form.save(commit=False)
            orcamento.cliente = cliente
            orcamento.veiculo = veiculo  # Defina o veículo associado diretamente aqui
            orcamento.save()

            total_orcamentos = OrcamentoNaoAprovado.objects.count()
            messages.success(request, 'Orçamento criado com sucesso!', extra_tags='success', fail_silently=True)
            return redirect('orcamentos_abertos')

    else:
        form = CriarOrcamentoForm()

    return render(request, 'pages/criar_orcamento.html', {'form': form, 'cliente': cliente, 'veiculo': veiculo})


def orcamentos_lista(request):
    clientes_com_orcamentos = Cliente.objects.filter(
        orcamentonaoaprovado__isnull=False
    ).annotate(
        num_orcamentos_nao_aprovados=Count('orcamentonaoaprovado')
    ).prefetch_related('veiculo', 'orcamentonaoaprovado').values(
        'id',
        'nome',
        'cpf',
        'telefone',
        'email',
        'veiculo__modelo',
        'veiculo__placa',
        'orcamentonaoaprovado__valor_estimado'
    )

    clientes_orcamentos = list(clientes_com_orcamentos)
    return render(request, 'pages/orcamentos_abertos.html', {'clientes': clientes_com_orcamentos})


def orcamentos_abertos(request):
    queryset = OrcamentoNaoAprovado.objects.filter(numero_orcamento__isnull=False)\
                                           .values(
                                               'id',
                                               'numero_orcamento',
                                               'cliente__nome',
                                               'cliente__cpf',
                                               'cliente__telefone',
                                               'cliente__email',
                                               'cliente__veiculo__marca',
                                               'cliente__veiculo__modelo',
                                               'cliente__veiculo__placa',
                                               'valor_estimado'
                                           )
    result_list = list(queryset)
    orcamentos_com_numero = []

    for item in result_list:
        orcamentos_com_numero.append(item)

    orcamentos_com_numero = list(queryset)

    return render(request, 'pages/orcamentos_abertos.html', {'orcamentos_com_numero': orcamentos_com_numero})


def excluir_orcamento(request, orcamento_id):
    orcamento = get_object_or_404(OrcamentoNaoAprovado, id=orcamento_id)
    orcamento.delete()

    # Redirecionar para a página de lista de orçamentos ou onde você preferir
    return redirect(reverse('orcamentos_abertos'))


# Função para aprovar orçamento e encaminhar para Ordem de Serviço em aberto
def os_aberta(request, orcamento_id):
    orcamento = OrcamentoNaoAprovado.objects.get(id=orcamento_id)

    ordem_aberta = orcamento.criar_ordem_servico_aberta()

    orcamento.delete()

    messages.success(request, 'Orçamento foi aprovado. Ordem de serviço iniciada! ', extra_tags='success', fail_silently=True)
    return redirect(reverse('ordem_servico'))

def ordem_servico(request):
    query_ordem = OrdemServicoAberta.objects.values('numero_ordem',
                                                    'data_ordem',
                                                    'servico',
                                                    'valor_estimado',
                                                    'cliente__nome',
                                                    'cliente__cpf',
                                                    'cliente__telefone',
                                                    'cliente__email',
                                                    'cliente__endereco',
                                                    'cliente__veiculo__marca',
                                                    'cliente__veiculo__modelo',
                                                    'cliente__veiculo__placa',
                                                    'cliente__veiculo__ano',
                                                    )
    result_query =[]

    for ordem in query_ordem:
        result_query.append(ordem)


    context = {
        'result_query': result_query,

    }
    return render(request, 'pages/os_aberta.html', context)


def concluir_ordem_servico(request, numero_ordem):
    ordem_aberta = OrdemServicoAberta.objects.get(numero_ordem=numero_ordem)

    # Transferir dados da ordem aberta para ordem concluída
    ordem_aberta.transfer_to_concluida()

    # Remover a ordem aberta após a transferência
    ordem_aberta.delete()

    messages.success(request, 'Ordem de serviço concluída e movida para histórico.', extra_tags='success', fail_silently=True)
    return redirect(reverse('ordens_concluidas'))  # Redirecionar de volta para a lista de ordens abertas


def ordens_concluidas(request):
    query_ordem_fechada = OrdemServicoConcluida.objects.values(
        'numero_ordem',
        'data_conclusao_ordem',
        'servico',
        'valor_total',
        'cliente__nome',
        'cliente__cpf',
        'cliente__telefone',
        'cliente__email',
        'cliente__endereco',
        'veiculo__marca',
        'veiculo__modelo',
        'veiculo__placa',
        'veiculo__ano',
    )


    result_query_fechada = list(query_ordem_fechada)

    context = {
        'result_query_fechada': result_query_fechada,
    }
    return render(request, 'pages/os_concluida.html', context)

def ordem_lista(request):
    query_ordem = OrdemServicoAberta.objects.values('numero_ordem',
                                                    'data_ordem',
                                                    'servico',
                                                    'valor_estimado',
                                                    'cliente__nome',
                                                    'cliente__cpf',
                                                    'cliente__telefone',
                                                    'cliente__email',
                                                    'cliente__endereco',
                                                    'cliente__veiculo__marca',
                                                    'cliente__veiculo__modelo',
                                                    'cliente__veiculo__placa',
                                                    'cliente__veiculo__ano',
                                                    )
    result_query = []

    for ordem in query_ordem:
        result_query.append(ordem)

    return JsonResponse({'result_query': result_query})

def ordens_concluidas_list(request):
    query_ordem_fechada = OrdemServicoConcluida.objects.values(
        'numero_ordem',
        'data_conclusao_ordem',
        'servico',
        'valor_total',
        'cliente__nome',
        'cliente__cpf',
        'cliente__telefone',
        'cliente__email',
        'cliente__endereco',
        'veiculo__marca',
        'veiculo__modelo',
        'veiculo__placa',
        'veiculo__ano',
    )
    quantidade_ordens = len(query_ordem_fechada)

    response_data = {
        'quantidade_ordens': quantidade_ordens,
        'query_result': list(query_ordem_fechada)
    }

    return JsonResponse(response_data)

def contar_veiculos_por_cpf(request, cpf):
    cliente = Cliente.objects.filter(cpf=cpf).first()
    if cliente:
        quantidade_veiculos = cliente.veiculo_set.count()
        return HttpResponse(f"Quantidade de veículos associados ao CPF {cpf}: {quantidade_veiculos}")
    else:
        return HttpResponse("CPF não encontrado ou sem veículos associados.")

def gerar_pdf_ordem(request, numero_ordem):
    # Obtenha os dados da ordem com base no número
    ordem = OrdemServicoConcluida.objects.get(numero_ordem=numero_ordem)

    # Carregue o template do PDF
    template_path = 'pages/ordem_pdf.html'
    template = get_template(template_path)
    context = {'ordem': ordem}

    # Renderize o template do PDF
    html = template.render(context)
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'filename="ordem_{numero_ordem}.pdf"'

    # Gere o PDF usando o xhtml2pdf
    pisa_status = pisa.CreatePDF(html, dest=response)
    if pisa_status.err:
        return HttpResponse('Erro ao gerar PDF', content_type='text/plain')
    return response